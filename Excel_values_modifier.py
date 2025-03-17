import openpyxl
import random
import os

# File name
file_name = "Журналы.xlsx"

# Check if the file exists
if not os.path.exists(file_name):
    print(f"Файл {file_name} не найден.")
    exit()

# Open the Excel file
try:
    wb = openpyxl.load_workbook(file_name, data_only=False)  # Keep formulas
    ws = wb.active
except Exception as e:
    print(f"Ошибка при открытии файла: {e}")
    exit()

# Find all unique names in row 6 starting with "Rp"
rp_columns = {}  # Dictionary {column: name}
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(row=6, column=col).value
    if isinstance(cell_value, str) and cell_value.startswith("Rp"):
        rp_columns[col] = cell_value

# Find rows containing X, Y, H 0 in column C
xyh_rows = {}  # Dictionary {row: type (X/Y/H 0)}
for row in range(1, ws.max_row + 1):
    cell_value = ws.cell(row=row, column=3).value
    if isinstance(cell_value, str) and cell_value.strip() in ["Х", "У", "Н 0", "Нn", "Нn+1"]:
        xyh_rows[row] = cell_value.strip()

# Modify values in found cells
for col in rp_columns:
    for row in xyh_rows:
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, (int, float)):  # Only numbers
            delta = round(random.uniform(0.001, 0.999), 3)  # Generate a random number
            new_value = round(cell.value + random.choice([-1, 1]) * delta, 3)  # +- change
            cell.value = new_value  # Write new value

# Save changes to the file
new_file_name = "Измененный_" + file_name
try:
    wb.save(new_file_name)
    print(f"Файл сохранен как: {new_file_name}")
except Exception as e:
    print(f"Ошибка при сохранении файла: {e}")
